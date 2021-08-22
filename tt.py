import openpyxl
import os
import glob
from pathlib import Path


posts = [
 
]
list_of_files = glob.glob('C:\\*\\*\\Downloads\\*.xlsx') # * means all if need specific format then *.csv
path = max(list_of_files, key=os.path.getctime)
print(path)
# path = "C:\\Users\\Raj\\Downloads\\Rough.xlsx"
wb_obj = openpyxl.load_workbook(path)

sheets = wb_obj.sheetnames

sheet_obj = wb_obj[sheets[1]]

count = 0
key=[]
value=[]

for i in sheet_obj.iter_cols(min_row=2, min_col=1, max_row=4, max_col=2):
    count = count + 1
    for cell in i:
        if count == 1:
            key.append(cell.value)
            #print(cell.value)
        else:
            value.append(cell.value)
    #print()

dict_room = {}
time = ['MWF9', 'MWF10', 'MWF11', 'MWF12', 'TT9', 'TT11', 'TT12', 'TT15']
courses = ['cs101', 'cs102', 'cs110', 'cs120', 'cs220', 'cs412', 'cs430', 'cs612', 'cs630' ]

rows_1,cols_1 = (5,7)

room_101 = [[0 for i in range(cols_1)] for j in range(rows_1)]

room_115 = [[0 for i in range(cols_1)] for j in range(rows_1)]

room_200 = [[0 for i in range(cols_1)] for j in range(rows_1)]

for i in range (0,len(key)):
    key_value = "{}".format(key[i])
    dict_room[key_value] = value[i]

print(dict_room)

def get_key(val):
    for key, value in dict_room.items():
         if val <= value:
             return key

def remove(string):
    return string.replace(" ", "")

def largest(arr,n):
    max = arr[0]
    for i in range(1, n):
        if arr[i] > max:
            max = arr[i]
    return max

def get_room_no(room_no):
    if room_no == '101':
        return room_101
    elif room_no == '200':
        return room_200
    else:
        return room_115

def allot_timetable(course_no,p,e):
    print("*******Processing******")
    room_no = get_key(e)
    if course_no == '101':
        course_no=101
    elif course_no == '630':
        course_no = 630
    elif course_no == '412':
        course_no = 412
    elif course_no == '612':
        course_no = 612
    else:
        if course_no == '501':
            course_no = 501
    #print("Allotted Room",room_no)
    room_allotted = get_room_no(room_no)
    pref = list(p.split(","))
    #print(pref)
    available_time_mwf = []
    available_time_tt= []
    day_count = 0
    time_count=9
    lec_count = 3
    for i in range(len(pref)):
        if pref[i].find('MWF') != -1 :
            time_available = pref[i]
            time_available = remove(time_available)
            time_available = time_available[slice(3,len(time_available)+1)]
            available_time_mwf.append(time_available)
        else:
            time_available = pref[i]
            time_available = remove(time_available)
            time_available = time_available[slice(2,len(time_available)+1)]
            available_time_tt.append(time_available)
            #print(time_available)
    #print(available_time_mwf)
    #print(available_time_tt)
    #print("Lecture Count: ",lec_count)
    for time in available_time_mwf:
        time = int(time) - 9
        for i in range(0,6,2):
            if room_allotted[i][time] == 0:
                #print("Alloted at time MWF",i,time)
                room_allotted[i][time] = str(course_no)
                lec_count -= 1
        if lec_count <= 0:
            break
    #print("Passed for Monday, Wednesday and Friday")
    #print(room_allotted)
    #print("Lecture Count: ",lec_count)
    if lec_count <= 0:
        print("Passed for all days")
        #print(room_allotted)
        #print("Lecture Count: ",lec_count)
        return
    else:
        for time in available_time_tt:
            time = int(time) - 9
            for i in range(1,5,2):
                if room_allotted[i][time] == 0:
                    print("Alloted at time for TT",i,time)
                    room_allotted[i][time] = str(course_no)
                    lec_count -= 1
                if lec_count <= 0:
                    break
        print("Passed for all days")
        #print(room_allotted)
        #print("Lecture Count: ",lec_count)
        return
    
                
            

    



    


#Give input here

input_obj = wb_obj[sheets[2]]

col=0
course_no=""
p=""
e=0



for row in input_obj.iter_rows(min_row=2, min_col=1, max_row=6, max_col=3):
    for cell in row:
        if col == 0:
            print("Checking for course",cell.value)
            course_no=str(cell.value)
        if col == 1:
            #print("Enrollment:",cell.value)
            e=cell.value
        if col == 2: 
            #print("Preferences:",cell.value)
            p = str(cell.value)
        col = col + 1
    #print(course_no)
    allot_timetable(course_no,p,e)
    print()
    col=0


print("Schedule for Room 101",room_101)

print("Schedule for Room 200",room_200)

print("Schedule for Room 115",room_115)