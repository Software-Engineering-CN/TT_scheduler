from flask import Flask, render_template, url_for, flash, redirect
from forms import RegistrationForm, LoginForm
import openpyxl
from pathlib import Path
app = Flask(__name__)
app.config['SECRET_KEY'] = '5791628bb0b13ce0c676dfde280ba245'

posts = [
 
]

path = "C:\\Users\Raj\\Documents\\software engineering subject\\Rough.xlsx"

wb_obj = openpyxl.load_workbook(path)

sheets = wb_obj.sheetnames

sheet_obj = wb_obj[sheets[1]]

count = 0
key=[]
value=[]

for i in sheet_obj.iter_cols(min_row=2, min_col=1, max_row=4, max_col=2):
    count = count + 1
    for cell in i:
        if count == 1:
            key.append(cell.value)
            #print(cell.value)
        else:
            value.append(cell.value)
    #print()

dict_room = {}
time = ['MWF9', 'MWF10', 'MWF11', 'MWF12', 'TT9', 'TT11', 'TT12', 'TT15']
courses = ['cs101', 'cs102', 'cs110', 'cs120', 'cs220', 'cs412', 'cs430', 'cs612', 'cs630' ]

rows_1,cols_1 = (5,7)

room_101 = [['0']*cols_1]*rows_1

room_115 = [['0']*cols_1]*rows_1

room_200 = [['0']*cols_1]*rows_1

for i in range (0,len(key)):
    key_value = "{}".format(key[i])
    dict_room[key_value] = value[i]

print(dict_room)

def get_key(val):
    for key, value in dict_room.items():
         if val <= value:
             return key

def remove(string):
    return string.replace(" ", "")

def largest(arr,n):
    max = arr[0]
    for i in range(1, n):
        if arr[i] > max:
            max = arr[i]
    return max

def allot_timetable(course_no,p,e):
    print("*******Processing******")
    room_no = get_key(e)
    if course_no == '101':
        course_no=101
    elif course_no == '630':
        course_no = 630
    elif course_no == '412':
        course_no = 412
    else:
        if course_no == '630':
            course_no = 630
    # print(course_no.type())
    print("Allotted Room",room_no)
    if room_no == 115 and course_no >= 600:
        pref = list(p.split(","))
        print(pref)
        available_time_mwf = []
        available_time_tt= []
        day_count = 0
        time_count=9
        for i in range(len(pref)):
            if pref[i].find('MWF') != -1 :
                time_available = pref[i]
                time_available = remove(time_available)
                time_available = time_available[slice(3,len(time_available)+1)]
                available_time_mwf.append(time_available)
            else:
                time_available = pref[i]
                time_available = remove(time_available)
                time_available = time_available[slice(2,len(time_available)+1)]
                available_time_tt.append(time_available)
                #print(time_available)
        print(available_time_mwf)
        print(available_time_tt)
        print("Checking for available time")
        for i in range(len(available_time_mwf)):
            if day_count == 0:
                row = 0
                col = int(largest(available_time_mwf,len(available_time_mwf)))
                print(col)
                col = 9 - col
                if room_115[row][col] == '0':
                    print("Alloting time",col,"on",row)
                    room_115[row][col] = course_no
            if day_count == 2:
                row = 2
                col = int(largest(available_time_mwf,len(available_time_mwf)))
                col = 9 - col
                print(col)
                if room_115[row][col] == '0':
                    print("Alloting time",col,"on",row)
                    room_115[row][col] = course_no
            if day_count == 4:
                row = 4
                col = int(largest(available_time_mwf,len(available_time_mwf)))
                col = 9 - col
                print(col)
                if room_115[row][col] == '0':
                    print("Alloting time",col,"on",row)
                    room_115[row][col] = course_no
            day_count += 2
        day_count=1
        for i in range(len(available_time_tt)):
            if day_count == 1:
                row = 1
                col = int(largest(available_time_tt,len(available_time_tt)))
                col = 9 - col
                print(col)
                if room_115[row][col] == '0':
                    print("Alloting time",col,"on",row)
                    room_115[row][col] = course_no
            if day_count == 3:
                row = 3
                col = int(largest(available_time_tt,len(available_time_tt)))
                col = 9 - col
                print(col)
                if room_115[row][col] == '0':
                    print("Alloting time",col,"on",row)
                    room_115[row][col] = course_no
        day_count += 1
        print(room_115)
    elif room_no == 200 and course_no < 600:
        pref = list(p.split(","))
        print(pref)
        available_time_mwf = []
        available_time_tt= []
        day_count = 0
        time_count=9
        for i in range(len(pref)):
            if pref[i].find('MWF') != -1 :
                time_available = pref[i]
                time_available = remove(time_available)
                time_available = time_available[slice(3,len(time_available)+1)]
                available_time_mwf.append(time_available)
            else:
                time_available = pref[i]
                time_available = remove(time_available)
                time_available = time_available[slice(2,len(time_available)+1)]
                available_time_tt.append(time_available)
                #print(time_available)
        print(available_time_mwf)
        print(available_time_tt)
        print("Checking for available time")
        for i in range(len(available_time_mwf)):
            if day_count == 0:
                row = 0
                col = int(largest(available_time_mwf,len(available_time_mwf)))
                print(col)
                col = 9 - col
                if room_200[row][col] == '0':
                    print("Alloting time",col,"on",row)
                    room_200[row][col] = course_no
            if day_count == 2:
                row = 2
                col = int(largest(available_time_mwf,len(available_time_mwf)))
                col = 9 - col
                print(col)
                if room_200[row][col] == '0':
                    print("Alloting time",col,"on",row)
                    room_200[row][col] = course_no
            if day_count == 4:
                row = 4
                col = int(largest(available_time_mwf,len(available_time_mwf)))
                col = 9 - col
                print(col)
                if room_200[row][col] == '0':
                    print("Alloting time",col,"on",row)
                    room_200[row][col] = course_no
            day_count += 2
        day_count=1
        for i in range(len(available_time_tt)):
            if day_count == 1:
                row = 1
                col = int(largest(available_time_tt,len(available_time_tt)))
                col = 9 - col
                print(col)
                if room_200[row][col] == '0':
                    print("Alloting time",col,"on",row)
                    room_200[row][col] = course_no
            if day_count == 3:
                row = 3
                col = int(largest(available_time_tt,len(available_time_tt)))
                col = 9 - col
                print(col)
                if room_200[row][col] == '0':
                    print("Alloting time",col,"on",row)
                    room_200[row][col] = course_no
        day_count += 1
        print(room_200)
    else:
        pref = list(p.split(","))
        print(pref)
        available_time_mwf = []
        available_time_tt= []
        day_count = 0
        time_count=9
        for i in range(len(pref)):
            if pref[i].find('MWF') != -1 :
                time_available = pref[i]
                time_available = remove(time_available)
                time_available = time_available[slice(3,len(time_available)+1)]
                available_time_mwf.append(time_available)
            else:
                time_available = pref[i]
                time_available = remove(time_available)
                time_available = time_available[slice(2,len(time_available)+1)]
                available_time_tt.append(time_available)
                #print(time_available)
        print(available_time_mwf)
        print(available_time_tt)
        print("Checking for available time")
        for i in range(len(available_time_mwf)):
            if day_count == 0:
                row = 0
                col = int(largest(available_time_mwf,len(available_time_mwf)))
                print(col)
                col = 9 - col
                if room_101[row][col] == '0':
                    print("Alloting time",col,"on",row)
                    room_101[row][col] = course_no
            if day_count == 2:
                row = 2
                col = int(largest(available_time_mwf,len(available_time_mwf)))
                col = 9 - col
                print(col)
                if room_101[row][col] == '0':
                    print("Alloting time",col,"on",row)
                    room_101[row][col] = course_no
            if day_count == 4:
                row = 4
                col = int(largest(available_time_mwf,len(available_time_mwf)))
                col = 9 - col
                print(col)
                if room_101[row][col] == '0':
                    print("Alloting time",col,"on",row)
                    room_101[row][col] = course_no
            day_count += 2
        day_count=1
        for i in range(len(available_time_tt)):
            if day_count == 1:
                row = 1
                col = int(largest(available_time_tt,len(available_time_tt)))
                col = 9 - col
                print(col)
                if room_101[row][col] == '0':
                    print("Alloting time",col,"on",row)
                    room_101[row][col] = course_no
            if day_count == 3:
                row = 3
                col = int(largest(available_time_tt,len(available_time_tt)))
                col = 9 - col
                print(col)
                if room_101[row][col] == '0':
                    print("Alloting time",col,"on",row)
                    room_101[row][col] = course_no
        day_count += 1
        print(room_101)
    



    



input_obj = wb_obj[sheets[2]]

col=0
course_no=""
p=""
e=0

for row in input_obj.iter_rows(min_row=2, min_col=1, max_row=5, max_col=3):
    for cell in row:
        if col == 0:
            print("Checking for course",cell.value)
            course_no=str(cell.value)
        if col == 1:
            #print("Enrollment:",cell.value)
            e=cell.value
        if col == 2: 
            #print("Preferences:",cell.value)
            p = str(cell.value)
        col = col + 1
    allot_timetable(course_no,p,e)
    print()
    col=0
    

@app.route("/")
@app.route("/home")
def home():
    return render_template('home.html', room=room_101)


@app.route("/about")
def about():
    return render_template('about.html', title='About')


@app.route("/register", methods=['GET', 'POST'])
def register():
    form = RegistrationForm()
    if form.validate_on_submit():
        flash(f'Account created for {form.username.data}!', 'success')
        return redirect(url_for('home'))
    return render_template('register.html', title='Register', form=form)


@app.route("/login", methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        if form.email.data == 'admin@blog.com' and form.password.data == 'password':
            flash('You have been logged in!', 'success')
            return redirect(url_for('home'))
        else:
            flash('Login Unsuccessful. Please check username and password', 'danger')
    return render_template('login.html', title='Login', form=form)


if __name__ == '__main__':
    app.run(debug=True)